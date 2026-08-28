"""
Export functionality for Surrey Social Value Platform
"""
import io
from datetime import datetime

from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.permissions import AllowAny

# Excel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# PDF
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER

from .models import VCSEResult


class ExcelExportView(APIView):
    permission_classes = [AllowAny]
    
    def get(self, request):
        wb = Workbook()
        ws = wb.active
        ws.title = "Projects"
        
        # Headers
        headers = ['Project', 'Organisation', 'Area', 'Volunteers', 'Hours', 'Total Value', 'Public Value', 'SROI']
        ws.append(headers)
        
        # Style headers
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1E3A5F", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        for project in VCSEResult.objects.all():
            ws.append([
                project.project_title,
                project.organisation_name,
                project.area_served,
                project.number_of_volunteers,
                project.annual_volunteer_hours,
                float(project.total_social_value),
                float(project.public_sector_value),
                float(project.sroi_ratio or 0)
            ])
        
        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            col_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 2, 40)
        
        # Save
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"surrey_projects_{datetime.now().strftime('%Y%m%d')}.xlsx"
        response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class PDFExportView(APIView):
    permission_classes = [AllowAny]
    
    def get(self, request):
        project_id = request.GET.get('project_id')
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
        
        styles = getSampleStyleSheet()
        elements = []
        
        # Title
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=20, textColor=colors.HexColor('#1E3A5F'), alignment=TA_CENTER)
        elements.append(Paragraph("Surrey Social Value Report", title_style))
        elements.append(Spacer(1, 0.3*inch))
        
        if project_id:
            # Single project report
            try:
                p = VCSEResult.objects.get(id=project_id)
                data = [
                    ['Field', 'Value'],
                    ['Project', p.project_title],
                    ['Organisation', p.organisation_name],
                    ['Area', p.area_served],
                    ['Volunteers', str(p.number_of_volunteers)],
                    ['Annual Hours', str(p.annual_volunteer_hours)],
                    ['Total Social Value', f"£{float(p.total_social_value):,.2f}"],
                    ['Public Sector Value', f"£{float(p.public_sector_value):,.2f}"],
                    ['SROI', f"{float(p.sroi_ratio or 0):.2f}"]
                ]
            except VCSEResult.DoesNotExist:
                data = [['Error', 'Project not found']]
        else:
            # Summary of all projects
            projects = VCSEResult.objects.all()
            total = sum(float(p.total_social_value) for p in projects)
            data = [
                ['Metric', 'Value'],
                ['Total Projects', str(len(projects))],
                ['Total Social Value', f"£{total:,.2f}"]
            ]
        
        # Create table
        table = Table(data, colWidths=[3*inch, 3*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A5F')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F8FAFC')),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#CBD5E1')),
            ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
        ]))
        elements.append(table)
        
        doc.build(elements)
        
        pdf = buffer.getvalue()
        buffer.close()
        
        filename = f"surrey_report_{datetime.now().strftime('%Y%m%d')}.pdf"
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response.write(pdf)
        return response